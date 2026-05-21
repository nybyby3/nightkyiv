#!/usr/bin/env python3
"""
NightKyiv 2 data extractor.
Reads all xlsx files from `Night Kyiv 2/` and produces:
  - venues.json (sport / entertainment / nature / culture / nightlife / food / hobby places)
  - recipes.json (cookbooks)
  - categories.json (the funnel taxonomy)
"""
import openpyxl, os, json, re, sys, unicodedata
from pathlib import Path
from collections import OrderedDict

SRC = Path("/sessions/zealous-kind-ride/mnt/Night Kyiv 2")
OUT = Path("/sessions/zealous-kind-ride/mnt/nightkyiv/src/data")
OUT.mkdir(parents=True, exist_ok=True)

def nfc(s): return unicodedata.normalize("NFC", s)
# macOS NFD-encoded -> NFC map
FILE_MAP = {nfc(p.stem): p for p in SRC.glob("*.xlsx")}

# ---------- TAXONOMY ----------
# Top-level buckets -> sub-categories -> source filename (without .xlsx)
TAXONOMY = OrderedDict([
    ("sport", {
        "uk": "Спорт", "en": "Sports", "ic": "🏋️", "color": "#4fc3f7",
        "subs": OrderedDict([
            ("martial_arts", {"uk":"Бойові мистецтва","en":"Martial Arts","ic":"🥋","files":[
                "Боксерські_клуби_Києва","Кікбоксинг_Києва","Тхеквондо_Києва","Карате_Києва",
                "Дзюдо_Києва","Боротьба_Києва","Фехтування_Києва","Капоейра_Києва",
                "Айкідо_Києва","Ушу_та_кунг-фу_Києва","Самбо_Києва","Муай-тай_Києва","MMA_Києва"]}),
            ("team", {"uk":"Командні","en":"Team Sports","ic":"⚽","files":[
                "Футбол_Києва","Баскетбол_Києва","Волейбол_Києва","Хокей_Києва",
                "Водне_поло_Києва","Бейсбол_і_софтбол_Києва","Гандбол_Києва",
                "Фрісбі_Києва","Регбі_Києва"]}),
            ("racket", {"uk":"Ракеткові","en":"Racket","ic":"🎾","files":[
                "Теніс_Києва","Настільний_теніс_Києва","Бадмінтон_Києва","Сквош_Києва","Падел_Києва"]}),
            ("water", {"uk":"Водні","en":"Water","ic":"🏊","files":[
                "Плавання_Києва","Дайвінг_Києва","Гребля_Києва",
                "Каное_каяки_сапы_Києва","Вейкбординг_та_водні_лижі_Києва"]}),
            ("winter", {"uk":"Зимові","en":"Winter","ic":"⛷️","files":[
                "Фігурне_катання_Києва","Гірські_лижі_та_сноубординг_Києва"]}),
            ("fitness", {"uk":"Фітнес","en":"Fitness","ic":"💪","files":[
                "Спортзали_Києва","Тяжка_атлетика_Києва","Армрестлінг_Києва","Пауерліфтинг_Києва",
                "Бодібілдинг_Києва","Стрітліфтинг_Києва","Воркаут_Києва","Гімнастика_Києва",
                "Акробатика_Києва","Йога_Києва","Пілатес_Києва","Танцювальний_фітнес_Києва",
                "Аквааеробіка_Києва","Степ-аеробіка_Києва","Стретчинг_Києва","Цигун_Тай_чи_Києва",
                "Кросфіт_Києва"]}),
            ("extreme", {"uk":"Екстрім","en":"Extreme","ic":"🤸","files":[
                "Скелелазіння_Києва","Паркур_Києва","Скейтбординг_Києва","Ролики_Києва",
                "BMX_Києва","Слеклайн_Києва","Даунхіл_Києва","Аеротруба_Києва",
                "Роуп-Джампінг_Києва","Зиплайн_Києва"]}),
            ("athletics", {"uk":"Легка атлетика","en":"Athletics","ic":"🏃","files":[
                "Легка_атлетика_Києва","Біг_Києва","Велоспорт_Києва"]}),
            ("aim_sports", {"uk":"Стрільба & гольф","en":"Aim sports","ic":"🎯","files":[
                "Стрільба_з_Лука_Києва","Гольф_Києва"]}),
            ("equestrian", {"uk":"Кінний спорт","en":"Equestrian","ic":"🐎","files":["Кінний_Спорт_Києва"]}),
            ("motor", {"uk":"Мотоспорт","en":"Motorsport","ic":"🏎️","files":[
                "Мотокрос_Києва","Дрифт_Києва","Ендуро_Києва","Квадроцикли_Києва","Картинг_Києва"]}),
        ])
    }),
    ("nature", {
        "uk":"Природа","en":"Nature","ic":"🌳","color":"#3ecf8e",
        "subs": OrderedDict([
            ("parks", {"uk":"Парки","en":"Parks","ic":"🌳","files":["Парки_Києва"]}),
            ("picnic_beach", {"uk":"Пляжі та пікніки","en":"Beaches & Picnics","ic":"🏖️","files":[
                "Пляжний_Відпочинок_Києва","Пікніки_Києва"]}),
            ("hiking", {"uk":"Походи","en":"Hiking","ic":"🥾","files":[
                "Пішохідні_походи_Києва","Кемпінг_Глемпинг_Києва","Бушкрафт_Києва"]}),
            ("nature_obs", {"uk":"Спостереження","en":"Observation","ic":"🔭","files":[
                "Бьордвотчинг_Києва","Астрономія_Києва","Фотоохота_Києва","Збирання_Києва"]}),
            ("fish_hunt", {"uk":"Риболовля та полювання","en":"Fishing & Hunting","ic":"🎣","files":[
                "Рибалка_Києва","Полювання_Києва"]}),
            ("zoo", {"uk":"Зоопарки","en":"Zoo / Aquariums","ic":"🦁","files":[
                "Зоопарки_ОКЕАНАРІУМи_Києва"]}),
            ("ecotour", {"uk":"Туризм","en":"Tourism","ic":"🗺️","files":[
                "Екотуризм_Києва","Фототуризм_Києва","Культурний_Туризм_Києва"]}),
        ])
    }),
    ("culture", {
        "uk":"Культура","en":"Culture","ic":"🎭","color":"#a259ff",
        "subs": OrderedDict([
            ("museums", {"uk":"Музеї","en":"Museums","ic":"🏛️","files":["Музеї_Києва"]}),
            ("galleries", {"uk":"Галереї","en":"Galleries","ic":"🖼️","files":["Галереї_Києва"]}),
            ("theatre", {"uk":"Театри","en":"Theatres","ic":"🎭","files":["Театральні_Постановки_Києва"]}),
            ("concerts", {"uk":"Концерти","en":"Concerts","ic":"🎵","files":["Концерти_Києва"]}),
            ("standup", {"uk":"Стенд-ап","en":"Stand-up","ic":"🎤","files":["Стенд-Ап_Києва"]}),
            ("libraries", {"uk":"Бібліотеки","en":"Libraries","ic":"📚","files":["Бібліотеки_Києва","Літературні_Клуби_Києва"]}),
            ("worship", {"uk":"Храми","en":"Worship","ic":"⛪","files":["Храми_Києва"]}),
            ("landmarks", {"uk":"Визначні місця","en":"Landmarks","ic":"📍","files":["Визначні_Місця_Києва"]}),
            ("art", {"uk":"Стрит-арт та живопис","en":"Street art","ic":"🎨","files":[
                "Граффіті_Києва","Живопис_Києва","Кіновиробництво_Києва"]}),
        ])
    }),
    ("nightlife", {
        "uk":"Нічне життя","en":"Nightlife","ic":"🌙","color":"#ff6bca",
        "subs": OrderedDict([
            ("bars", {"uk":"Бари","en":"Bars","ic":"🍸","files":["Бари_Києва"]}),
            ("clubs", {"uk":"Нічні клуби","en":"Clubs","ic":"🪩","files":["Нічні_клуби_Києва"]}),
            ("hookah", {"uk":"Кальянні","en":"Hookah","ic":"💨","files":["Кальянні_Києва"]}),
            ("wine", {"uk":"Винні дегустації","en":"Wine tastings","ic":"🍷","files":["Винні-Дегустації_Києва"]}),
        ])
    }),
    ("food", {
        "uk":"Їжа","en":"Food","ic":"🍽️","color":"#ff6b6b",
        "subs": OrderedDict([
            ("restaurants", {"uk":"Ресторани","en":"Restaurants","ic":"🍽️","files":["Ресторани_Києва"]}),
        ])
    }),
    ("entertainment", {
        "uk":"Розваги","en":"Entertainment","ic":"🎲","color":"#ff9f43",
        "subs": OrderedDict([
            ("bowling_darts", {"uk":"Боулінг & дартс","en":"Bowling & darts","ic":"🎳","files":["Боулінг_Києва","Дартс_Києва"]}),
            ("quests", {"uk":"Квести","en":"Escape rooms","ic":"🗝️","files":["Квести_Києва"]}),
            ("battle", {"uk":"Пейнтбол & лазертаг","en":"Paintball","ic":"🎯","files":[
                "Пейнтбол_Києва","Лазертаг_Києва","Страйкбол_Києва"]}),
            ("arcade", {"uk":"Аркади","en":"Arcades","ic":"🕹️","files":["Аркади_Києва","PC-Гейминг_Києва"]}),
            ("board", {"uk":"Настільні ігри & шахи","en":"Board games","ic":"♟️","files":[
                "Клуби-Настільних-Ігор_Києва","Шахи_Києва"]}),
            ("trampoline", {"uk":"Батути","en":"Trampolines","ic":"🤸","files":["Батути_Києва"]}),
            ("masterclass", {"uk":"Майстер-класи","en":"Masterclasses","ic":"🎓","files":["Майстер-Класи_Києва"]}),
            ("urban", {"uk":"Урбан-експлоринг","en":"Urban exploring","ic":"🏚️","files":["Урбанексплоринг_Києва"]}),
            ("kids", {"uk":"Для дітей","en":"For kids","ic":"🧒","files":["Розваги_Для_Дітей_Києва"]}),
        ])
    }),
    ("dance", {
        "uk":"Танці","en":"Dance","ic":"💃","color":"#36d1c4",
        "subs": OrderedDict([
            ("ballroom", {"uk":"Бальні","en":"Ballroom","ic":"💃","files":["Бальні_Танці_Києва"]}),
            ("latin", {"uk":"Латино","en":"Latin","ic":"🪇","files":["Латинські_Танці_Києва","Танго_Києва"]}),
            ("street", {"uk":"Стріт","en":"Street","ic":"🕺","files":["Хіп-Хоп_Києва","Брейк-Данс_Києва"]}),
            ("folk", {"uk":"Народні","en":"Folk","ic":"🥁","files":["Народні_Танці_Києва","Фламенко_Києва","Свінг_Києва"]}),
        ])
    }),
    ("hobby", {
        "uk":"Хобі вдома","en":"Hobbies","ic":"🎨","color":"#f7b731",
        "subs": OrderedDict([
            ("music", {"uk":"Музика","en":"Music","ic":"🎸","files":[
                "Гітара_Києва","Укулеле_Києва","Піаніно_Києва","Скрипка_Києва",
                "Барабани_Києва","Флейта_Києва","Саксофон_Києва","Диджеїнг_Києва"]}),
            ("crafts", {"uk":"Ремесла","en":"Crafts","ic":"🧵","files":[
                "Рисування_Києва","Ліплення_Києва","Шиття_Києва","Декупаж_Києва",
                "Миловаріння_Києва","Свічки_Києва","Ювелірна_справа_Києва",
                "Деревообробка_Києва","Розпис_по_склу_Києва","Оригамі_Києва",
                "Калліграфія_Києва","Скрапбукінг_Києва"]}),
            ("garden", {"uk":"Сад та рослини","en":"Garden","ic":"🌱","files":[
                "Кімнатні_Рослини_Києва","Город_на_Балконі_Києва","Город_Києва",
                "Флористика_Києва","Бонсай_Києва","Мікрозелень_Києва","Гідропоніка_Києва"]}),
            ("volunteer", {"uk":"Волонтерство","en":"Volunteer","ic":"🤝","files":["Волонтерство_Києва"]}),
        ])
    }),
])

# Cookbook taxonomy
COOK_SUBS = OrderedDict([
    ("bake", {"uk":"Випічка та хліб","en":"Baking","ic":"🍞","files":[
        "Кулинарная_книга_Выпечка_хлеба","Кулинарная_книга_Кондитерское_дело","Кулинарная_книга_Десерты"]}),
    ("meat", {"uk":"М'ясо та гриль","en":"Meat & grill","ic":"🥩","files":[
        "Кулинарная_книга_Гриль_и_BBQ","Кулинарная_книга_Копчение"]}),
    ("world", {"uk":"Кухні світу","en":"World cuisines","ic":"🌍","files":[
        "Кулинарная_книга_Азиатская_кухня","Кулинарная_книга_Веганская_кухня",
        "Кулинарная_книга_Итальянская_кухня","Кулинарная_книга_Французская_кухня",
        "Кулинарная_книга_Суши_и_японская","Кулинарная_книга_Экзотическая_кухня",
        "Кулинарная_книга_Молекулярная_кухня","Кулинарная_книга_Сыроедение"]}),
    ("drinks", {"uk":"Напої","en":"Drinks","ic":"☕","files":[
        "Кулинарная_книга_Кофе","Кулинарная_книга_Чайная_церемония",
        "Кулинарная_книга_Коктейли","Кулинарная_книга_Домашнее_пивоварение"]}),
    ("preserves", {"uk":"Соуси та заготовки","en":"Preserves & sauces","ic":"🥫","files":[
        "Кулинарная_книга_Заготовки","Кулинарная_книга_Соусы","Кулинарная_книга_Специи"]}),
    ("chocolate", {"uk":"Шоколад та солодке","en":"Sweets","ic":"🍫","files":[
        "Кулинарная_книга_Шоколад","Кулинарная_книга_Сыроделие"]}),
    ("plating", {"uk":"Подача та стайлінг","en":"Plating","ic":"🍱","files":[
        "Кулинарная_книга_Декор_блюд","Кулинарная_книга_Фуд-стайлинг"]}),
])

# ---------- HELPERS ----------
def clean(v):
    if v is None: return ""
    s = str(v).strip()
    return "" if s in ("—", "-", "None") else s

def is_section_header(row):
    """Detect rows that are visual section headers (only first cell filled)."""
    nn = [c for c in row if c is not None and str(c).strip() != ""]
    return len(nn) <= 1

def parse_rating(v):
    if v is None: return None
    try:
        if isinstance(v, (int, float)):
            return float(v) if 0 < v <= 5 else None
        m = re.search(r"(\d+[.,]\d+)", str(v))
        if m: return float(m.group(1).replace(",", "."))
    except: pass
    return None

def parse_int(v):
    if v is None: return None
    try:
        if isinstance(v, (int, float)): return int(v)
        m = re.search(r"(\d+)", str(v))
        if m: return int(m.group(1))
    except: pass
    return None

def parse_price(s):
    """Extract price range like '300-450 грн', '$$', '$$$' etc."""
    if not s: return None
    s = str(s)
    # Try numeric range
    m = re.search(r"(\d{2,5})\s*[-–—]\s*(\d{2,5})", s)
    if m: return {"min": int(m.group(1)), "max": int(m.group(2))}
    m = re.search(r"(\d{2,5})", s)
    if m: return {"min": int(m.group(1)), "max": int(m.group(1))}
    # $$$ style
    if "$" in s:
        return {"tier": s.count("$")}
    if re.search(r"free|безк|FREE", s, re.I):
        return {"min": 0, "max": 0}
    return None

def normalize_venue_row(row, header_map, file_cat, file_sub):
    """Map a raw row dict (header->value) into a unified venue."""
    def g(*keys):
        for k in keys:
            for hk, hv in header_map.items():
                if k.lower() in hk.lower():
                    val = clean(row.get(hv))
                    if val: return val
        return ""
    name = g("Назва", "Название", "клубу")
    if not name or is_section_header(list(row.values())): return None
    return {
        "name": name,
        "district": g("Район"),
        "address": g("Адреса", "Адрес"),
        "phone": g("Телефон"),
        "site": g("Сайт", "Соцмережі", "Сайт/Соц", "Інстаграм"),
        "rating": parse_rating(g("Рейтинг")),
        "reviews": parse_int(g("Відгуків", "К-сть відгуків")),
        "hours": g("Розклад", "Години", "Графік"),
        "price": parse_price(g("Ціни", "Чек", "Ціна", "Платно") or ""),
        "format": g("Формат", "Тип", "Спеціалізація", "Стиль", "Кухня", "Меню"),
        "level": g("Рівень"),
        "notes": g("Примітки", "Заметки"),
        "category": file_cat,
        "subcategory": file_sub,
        "flagship": bool(re.search(r"🏆|ФЛАГМАН|FLAGSHIP", g("Назва","Название","Примітки","Заметки"), re.I)),
    }

def parse_venue_file(path, file_cat, file_sub):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    venues = []
    for sn in wb.sheetnames:
        if sn.lower() in ("інформація", "информация", "гід", "guide"): continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: continue
        # find header row (first row containing 'Назва' or 'Название')
        header_idx = 0
        for i, r in enumerate(rows[:5]):
            if any(c and ("назва" in str(c).lower() or "название" in str(c).lower()) for c in r):
                header_idx = i; break
        header = rows[header_idx]
        header_map = {str(h).strip(): i for i, h in enumerate(header) if h}
        for r in rows[header_idx+1:]:
            row_dict = dict(zip(range(len(r)), r))
            v = normalize_venue_row(row_dict, {k: header_map[k] for k in header_map}, file_cat, file_sub)
            if v: venues.append(v)
    wb.close()
    return venues

def parse_recipe_file(path, sub_id):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    recipes = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        # Find header row
        header_idx = None
        for i, r in enumerate(rows[:10]):
            if any(c and "название" in str(c).lower() for c in r):
                header_idx = i; break
        if header_idx is None: continue
        header = [str(h).strip() if h else "" for h in rows[header_idx]]
        for r in rows[header_idx+1:]:
            if not r or not r[1]: continue
            d = {h: clean(v) for h, v in zip(header, r) if h}
            if not d.get("Название"): continue
            recipes.append({
                "name": d.get("Название",""),
                "country": d.get("Страна") or d.get("Страна / Регион") or d.get("Регион") or d.get("Кухня",""),
                "difficulty": d.get("Сложность",""),
                "category": d.get("Категория") or d.get("Тип") or d.get("Стиль") or d.get("Тема",""),
                "time": d.get("Время приготовления") or d.get("Общее время") or d.get("Время",""),
                "servings": d.get("Порций") or d.get("Выход",""),
                "ingredients": d.get("Ингредиенты") or d.get("Что нужно (ингредиенты)") or d.get("Ингредиенты (на 20 л)",""),
                "equipment": d.get("Кухонный инвентарь") or d.get("Оборудование") or d.get("Инвентарь") or d.get("Инструменты") or d.get("Барный инвентарь") or d.get("Чайная утварь") or d.get("Инвентарь и тара",""),
                "steps": d.get("Пошаговое приготовление") or d.get("Пошаговый процесс") or d.get("Подробная инструкция") or d.get("Пошаговое выполнение") or d.get("Пошаговое приготовление и церемония",""),
                "tips": d.get("Советы повара") or d.get("Советы шефа") or d.get("Советы шеф-повара") or d.get("Советы кондитера") or d.get("Советы хозяйки") or d.get("Советы пивовара") or d.get("Советы коптильщика") or d.get("Советы бариста") or d.get("Советы бармена") or d.get("Советы соусье") or d.get("Советы сыродела") or d.get("Советы шоколатье") or d.get("Советы чайного мастера") or d.get("Советы") or d.get("Совет профи",""),
                "subcategory": sub_id,
            })
    wb.close()
    return recipes

# ---------- EXTRACTION ----------
all_venues = []
missing = []
seen_files = set()

for cat_id, cat in TAXONOMY.items():
    for sub_id, sub in cat["subs"].items():
        for fbase in sub["files"]:
            key = nfc(fbase)
            seen_files.add(key)
            fpath = FILE_MAP.get(key)
            if fpath is None:
                missing.append(fbase); continue
            try:
                vs = parse_venue_file(fpath, cat_id, sub_id)
                for v in vs:
                    v["source"] = fbase
                all_venues.extend(vs)
            except Exception as e:
                print(f"ERR {fbase}: {e}", file=sys.stderr)

all_recipes = []
for sub_id, sub in COOK_SUBS.items():
    for fbase in sub["files"]:
        key = nfc(fbase)
        seen_files.add(key)
        fpath = FILE_MAP.get(key)
        if fpath is None:
            missing.append(fbase); continue
        try:
            rs = parse_recipe_file(fpath, sub_id)
            for r in rs: r["source"] = fbase
            all_recipes.extend(rs)
        except Exception as e:
            print(f"ERR recipe {fbase}: {e}", file=sys.stderr)

# Discover files not in taxonomy
all_files = set(FILE_MAP.keys())
orphans = sorted(all_files - seen_files)

# Stats
print(f"Venues:  {len(all_venues)}")
print(f"Recipes: {len(all_recipes)}")
print(f"Missing files (in taxonomy but no xlsx): {missing}")
print(f"Orphan files (xlsx not in taxonomy):    {orphans}")

# Build category tree for the app
cat_tree = []
for cid, c in TAXONOMY.items():
    cat_tree.append({
        "id": cid, "uk": c["uk"], "en": c["en"], "ic": c["ic"], "color": c["color"],
        "subs": [{"id": sid, "uk": s["uk"], "en": s["en"], "ic": s["ic"]} for sid, s in c["subs"].items()]
    })

cook_tree = [{"id": sid, "uk": s["uk"], "en": s["en"], "ic": s["ic"]} for sid, s in COOK_SUBS.items()]

# Assign stable IDs
for i, v in enumerate(all_venues): v["id"] = f"v{i+1}"
for i, r in enumerate(all_recipes): r["id"] = f"r{i+1}"

# Write
(OUT / "venues.json").write_text(json.dumps(all_venues, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
(OUT / "recipes.json").write_text(json.dumps(all_recipes, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
(OUT / "categories.json").write_text(json.dumps({"venues": cat_tree, "recipes": cook_tree}, ensure_ascii=False, indent=2), encoding="utf-8")

print(f"\nWrote to {OUT}")
print(f"  venues.json:      {(OUT/'venues.json').stat().st_size//1024} KB")
print(f"  recipes.json:     {(OUT/'recipes.json').stat().st_size//1024} KB")
print(f"  categories.json:  {(OUT/'categories.json').stat().st_size//1024} KB")

# Sample
print("\nSample venue:")
print(json.dumps(all_venues[0], ensure_ascii=False, indent=2))
print("\nSample recipe:")
if all_recipes:
    s = all_recipes[0].copy()
    for k in ("steps","ingredients","tips"):
        if s.get(k): s[k] = s[k][:120] + "..." if len(s[k]) > 120 else s[k]
    print(json.dumps(s, ensure_ascii=False, indent=2))
